"""
ГЛАВНЫЙ МОДУЛЬ ПРИЛОЖЕНИЯ ДЛЯ СРАВНЕНИЯ EXCEL-ТАБЕЛЕЙ
Автор: Ваше имя
Версия: 1.1
"""

# Импорт необходимых библиотек
import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QFileDialog, QLabel,
                             QProgressBar, QTextEdit, QMessageBox, QComboBox,
                             QSpinBox, QToolButton)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QIcon


class ExcelComparator(QMainWindow):
    """Главное окно приложения, наследующее QMainWindow"""

    def __init__(self):
        super().__init__()
        self.file1_path = None  # Путь к первому файлу
        self.file2_path = None  # Путь ко второму файлу
        self.current_worker = None  # Экземпляр рабочего потока
        self.selected_month = 1  # Выбранный месяц
        self.selected_year = 2025  # Выбранный год
        self.init_ui()  # Инициализация интерфейса

    def init_ui(self):
        """Инициализация пользовательского интерфейса"""
        # Основные настройки окна
        self.setWindowTitle("Сравнение Excel-файлов")
        self.setMinimumSize(800, 600)

        # Центральный виджет и основной layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Инициализация компонентов интерфейса
        self.init_date_selector(main_layout)  # Выбор даты
        self.init_file_selectors(main_layout)  # Выбор файлов
        self.init_controls(main_layout)  # Управляющие кнопки
        self.init_progress_log(main_layout)  # Прогресс и лог

        # Статусная строка внизу окна
        self.statusBar().showMessage('Готово к работе')

    def init_date_selector(self, layout):
        """Инициализация выбора месяца и года"""
        date_layout = QHBoxLayout()

        # Выпадающий список для месяцев
        self.month_combo = QComboBox()
        months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                  'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(0)  # Установка января по умолчанию

        # Спинбокс для выбора года
        self.year_spin = QSpinBox()
        self.year_spin.setRange(2000, 2100)  # Диапазон допустимых годов
        self.year_spin.setValue(2025)  # Значение по умолчанию

        # Добавление элементов в layout
        date_layout.addWidget(QLabel("Месяц табеля:"))
        date_layout.addWidget(self.month_combo)
        date_layout.addWidget(QLabel("Год:"))
        date_layout.addWidget(self.year_spin)
        layout.addLayout(date_layout)

    def init_file_selectors(self, layout):
        """Инициализация компонентов выбора файлов"""
        # Layout для первого файла
        file1_layout = QHBoxLayout()
        self.btn_file1 = QPushButton("Выбрать базовый файл", self)
        self.btn_file1.clicked.connect(lambda: self.select_file(1))
        self.label_file1 = QLabel("Файл не выбран")
        file1_layout.addWidget(self.btn_file1)
        file1_layout.addWidget(self.label_file1)

        # Layout для второго файла
        file2_layout = QHBoxLayout()
        self.btn_file2 = QPushButton("Выбрать файл для сравнения", self)
        self.btn_file2.clicked.connect(lambda: self.select_file(2))
        self.label_file2 = QLabel("Файл не выбран")
        file2_layout.addWidget(self.btn_file2)
        file2_layout.addWidget(self.label_file2)

        # Добавление в основной layout
        layout.addLayout(file1_layout)
        layout.addLayout(file2_layout)

    def init_controls(self, layout):
        """Инициализация элементов управления"""
        control_layout = QHBoxLayout()

        # Кнопка "О программе" с меню
        self.btn_about = QToolButton(self)
        self.btn_about.setText("ℹ️")
        self.btn_about.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.btn_about.clicked.connect(self.show_about_dialog)

        # Основная кнопка запуска сравнения
        self.btn_compare = QPushButton("Сравнить файлы", self)
        self.btn_compare.clicked.connect(self.start_comparison)

        # Кнопка прерывания процесса
        self.btn_abort = QPushButton("Прервать", self)
        self.btn_abort.clicked.connect(self.abort_processing)
        self.btn_abort.setEnabled(False)  # Изначально неактивна

        # Компоновка элементов управления
        control_layout.addWidget(self.btn_compare)
        control_layout.addWidget(self.btn_abort)
        control_layout.addWidget(self.btn_about)
        layout.addLayout(control_layout)

    def init_progress_log(self, layout):
        """Инициализация элементов отображения прогресса"""
        # Прогресс-бар
        self.progress = QProgressBar(self)
        self.progress.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.progress)

        # Текстовое поле для лога
        self.log = QTextEdit()
        self.log.setReadOnly(True)  # Режим только для чтения
        layout.addWidget(self.log)

    def show_about_dialog(self):
        """Отображение диалогового окна с информацией о программе"""
        about_text = """<b>Сравнение табелей учета рабочего времени</b><br><br>
        Версия 1.1<br>
        Автор: VaSeBa<br><br>
        Функционал:<br>
        - Сравнение двух Excel-файлов табелей<br>
        - Выделение различий цветом<br>
        - Генерация отчетов с различными категориями расхождений<br>
        """
        QMessageBox.information(self, "О программе", about_text)

    def select_file(self, file_num):
        """Выбор файла через диалоговое окно"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            "",
            "Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)"
        )

        if file_path:
            try:
                # Проверка читаемости файла
                pd.read_excel(file_path, nrows=1)
                if file_num == 1:
                    self.file1_path = file_path
                    self.label_file1.setText(os.path.basename(file_path))
                else:
                    self.file2_path = file_path
                    self.label_file2.setText(os.path.basename(file_path))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка",
                                     f"Невозможно прочитать файл:\n{str(e)}")

    def start_comparison(self):
        """Запуск процесса сравнения файлов"""
        # Получение выбранных дат
        self.selected_month = self.month_combo.currentIndex() + 1
        self.selected_year = self.year_spin.value()

        # Проверка выбранных файлов
        if not all([self.file1_path, self.file2_path]):
            QMessageBox.warning(self, "Ошибка", "Необходимо выбрать оба файла!")
            return

        if self.file1_path == self.file2_path:
            QMessageBox.warning(self, "Ошибка", "Файлы должны быть разными!")
            return

        # Создание и настройка рабочего потока
        self.current_worker = CompareWorker(
            self.file1_path,
            self.file2_path,
            self.selected_month,
            self.selected_year
        )

        # Подключение сигналов потока
        self.current_worker.progress_updated.connect(self.update_progress)
        self.current_worker.message_received.connect(self.log_message)
        self.current_worker.finished.connect(self.on_completion)
        self.current_worker.error_occurred.connect(self.show_error)

        # Запуск потока
        self.current_worker.start()
        self.toggle_controls(False)

    def abort_processing(self):
        """Прерывание выполнения операции"""
        if self.current_worker and self.current_worker.isRunning():
            self.current_worker.terminate()
            self.log_message("Процесс прерван пользователем")
            self.toggle_controls(True)

    def toggle_controls(self, enable):
        """Переключение состояния элементов управления"""
        self.btn_compare.setEnabled(enable)
        self.btn_abort.setEnabled(not enable)
        self.progress.setValue(0)
        self.log.clear()

    def update_progress(self, value):
        """Обновление значения прогресс-бара"""
        self.progress.setValue(value)

    def log_message(self, message):
        """Добавление сообщения в лог с временной меткой"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log.append(f"[{timestamp}] {message}")

    def show_error(self, message):
        """Обработка и отображение ошибок"""
        QMessageBox.critical(self, "Критическая ошибка", message)
        self.toggle_controls(True)

    def on_completion(self, output_path):
        """Действия по завершении обработки"""
        self.toggle_controls(True)
        if output_path:
            QMessageBox.information(
                self,
                "Готово",
                f"Результаты сравнения сохранены в:\n{output_path}"
            )
            self.progress.setValue(100)


class CompareWorker(QThread):
    """Класс-поток для выполнения сравнения в фоновом режиме"""

    # Объявление сигналов для взаимодействия с GUI
    progress_updated = pyqtSignal(int)  # Прогресс выполнения
    message_received = pyqtSignal(str)  # Сообщения в лог
    finished = pyqtSignal(str)  # Завершение работы
    error_occurred = pyqtSignal(str)  # Ошибки при выполнении

    def __init__(self, file1_path, file2_path, month, year):
        super().__init__()
        # Инициализация параметров
        self.file1_path = file1_path  # Путь к первому файлу
        self.file2_path = file2_path  # Путь ко второму файлу
        self.month = month  # Выбранный месяц
        self.year = year  # Выбранный год
        self._is_running = True  # Флаг выполнения потока

    def run(self):
        """Основная логика сравнения файлов"""
        try:
            self.message_received.emit("Инициализация процесса сравнения...")

            # Загрузка данных из файлов
            df1 = self.load_data(self.file1_path)
            df2 = self.load_data(self.file2_path)

            # Проверка структуры данных
            self.validate_data(df1, df2)

            # Объединение данных и обработка различий
            merged_df = self.merge_dataframes(df1, df2)
            report_data, highlight_info = self.process_differences(merged_df)

            # Генерация итогового отчета
            output_path = self.generate_report(report_data, highlight_info)
            self.finished.emit(output_path)

        except Exception as e:
            self.error_occurred.emit(f"Ошибка обработки: {str(e)}")
        finally:
            self._is_running = False

    def load_data(self, file_path):
        """Загрузка данных из Excel-файла"""
        self.message_received.emit(f"Загрузка {os.path.basename(file_path)}...")
        try:
            # Чтение файла с настройками
            df = pd.read_excel(
                file_path,
                dtype=str,  # Чтение всех данных как строки
                keep_default_na=False,  # Игнорирование NaN значений
                engine='openpyxl'  # Используемый движок
            )
            df.columns = df.columns.astype(str)  # Приведение названий колонок к строке
            return df
        except Exception as e:
            raise ValueError(f"Ошибка чтения файла {file_path}: {str(e)}")

    def validate_data(self, df1, df2):
        """Проверка структуры данных в файлах"""
        required_columns = {'id', 'ФИО', 'должность'} | \
                           {str(i) for i in range(1, 32)}  # Колонки дней

        for df, num in zip([df1, df2], ['первом', 'втором']):
            missing = required_columns - set(df.columns)
            if missing:
                raise ValueError(
                    f"В {num} файле отсутствуют колонки: {', '.join(missing)}")

    def merge_dataframes(self, df1, df2):
        """Объединение данных из двух DataFrame"""
        self.message_received.emit("Сопоставление данных...")
        # Добавление индексов для отслеживания позиций
        df1['original_index'] = df1.index
        df2['original_index'] = df2.index

        # Внешнее объединение по полю id
        merged = pd.merge(
            df1, df2,
            on='id',
            suffixes=('_base', '_compare'),
            how='outer',
            indicator=True)

        # Сохранение отсутствующих записей
        self.missing_in_base = merged[merged['_merge'] == 'right_only'][['id', 'ФИО_compare']]
        self.missing_in_compare = merged[merged['_merge'] == 'left_only'][['id', 'ФИО_base']]

        return merged[merged['_merge'] == 'both']

    def process_differences(self, merged_df):
        """Обработка и классификация расхождений"""
        self.message_received.emit("Поиск различий...")
        vv_diff = []  # Различия с меткой "ВВ"
        dp_diff = []  # Различия с меткой "ДП"
        other_diff = []  # Прочие различия
        highlight_info = {'base': [], 'compare': []}  # Ячейки для подсветки
        total_rows = len(merged_df)
        day_columns = [str(i) for i in range(1, 32)]  # Колонки дней (1-31)

        for idx, (_, row) in enumerate(merged_df.iterrows()):
            if not self._is_running:  # Проверка флага прерывания
                break

            # Получение индексов строк в исходных файлах
            base_index = row['original_index_base'] + 2  # +2 для Excel строк
            compare_index = row['original_index_compare'] + 2

            # Проверка различий по дням
            for day in day_columns:
                base_val = str(row[f'{day}_base']).strip()
                compare_val = str(row[f'{day}_compare']).strip()

                if base_val != compare_val:
                    # Формирование даты для отчета
                    try:
                        date_str = f"{int(day):02d}.{self.month:02d}.{self.year}"
                    except:
                        date_str = day

                    # Создание записи для отчета
                    entry = [
                        row['id'],
                        row['ФИО_base'],
                        date_str,
                        base_val,
                        compare_val
                    ]

                    # Классификация различий
                    if "ВВ" in (base_val, compare_val):
                        vv_diff.append(entry)
                    elif "ДП" in (base_val, compare_val):
                        dp_diff.append(entry)
                    else:
                        other_diff.append(entry)

                    # Сохранение позиций для подсветки
                    base_col = merged_df.columns.get_loc(f'{day}_base') + 1
                    compare_col = merged_df.columns.get_loc(f'{day}_compare') + 1
                    highlight_info['base'].append((base_index, base_col))
                    highlight_info['compare'].append((compare_index, compare_col))

            # Обновление прогресса выполнения
            progress = int(20 + 70 * (idx + 1) / total_rows)
            self.progress_updated.emit(progress)

        # Обработка отсутствующих записей
        missing_data = []
        for _, row in self.missing_in_base.iterrows():
            missing_data.append([row['id'], row['ФИО_compare'], "Отсутствует в БАЗОВОМ файле"])

        for _, row in self.missing_in_compare.iterrows():
            missing_data.append([row['id'], row['ФИО_base'], "Отсутствует в ФАЙЛЕ СРАВНЕНИЯ"])

        return (
            {'vv': vv_diff, 'dp': dp_diff, 'other': other_diff, 'missing': missing_data},
            highlight_info
        )

    def generate_report(self, report_data, highlight_info):
        """Генерация итогового отчета"""
        self.message_received.emit("Формирование отчетов...")
        try:
            # Подсветка различий в исходных файлах
            self.highlight_differences(self.file1_path, highlight_info['base'])
            self.highlight_differences(self.file2_path, highlight_info['compare'])

            # Создание файла отчета
            output_path = self.create_report_file(report_data)
            return output_path
        except Exception as e:
            raise ValueError(f"Ошибка создания отчетов: {str(e)}")

    def highlight_differences(self, file_path, cells):
        """Подсветка ячеек с различиями в файле"""
        if not cells:
            return

        # Создание объекта книги и листа
        wb = load_workbook(file_path)
        ws = wb.active

        # Настройка стиля подсветки
        yellow_fill = PatternFill(
            start_color='FFFF00',
            end_color='FFFF00',
            fill_type='solid'
        )

        # Применение стиля к ячейкам
        for row, col in cells:
            ws.cell(row=row, column=col).fill = yellow_fill

        # Сохранение изменений
        wb.save(file_path)
        wb.close()

    def create_report_file(self, report_data):
        """Создание файла отчета с несколькими листами"""
        # Формирование пути для сохранения
        output_dir = os.path.dirname(self.file1_path)
        base_name = os.path.splitext(os.path.basename(self.file1_path))[0]
        output_name = f"Сравнение_{base_name}.xlsx"
        output_path = os.path.join(output_dir, output_name)

        # Загрузка базового файла
        wb = load_workbook(self.file1_path)

        # Удаление существующих листов отчетов
        for sheet in ['ВВ', 'ДП', 'Остальные', 'Отсутствующие']:
            if sheet in wb.sheetnames:
                del wb[sheet]

        # Конфигурация листов отчета
        sheets_config = {
            'ВВ': {
                'data': report_data['vv'],
                'color': 'FF0000',  # Красный
                'header': f'Различия с отметками ВВ ({self.month:02d}.{self.year})',
                'columns': ['ID', 'ФИО', 'Дата', 'Базовый файл', 'Файл сравнения']
            },
            'ДП': {
                'data': report_data['dp'],
                'color': '0000FF',  # Синий
                'header': f'Различия с отметками ДП ({self.month:02d}.{self.year})',
                'columns': ['ID', 'ФИО', 'Дата', 'Базовый файл', 'Файл сравнения']
            },
            'Остальные': {
                'data': report_data['other'],
                'color': '008000',  # Зеленый
                'header': f'Прочие различия ({self.month:02d}.{self.year})',
                'columns': ['ID', 'ФИО', 'Дата', 'Базовый файл', 'Файл сравнения']
            },
            'Отсутствующие': {
                'data': report_data['missing'],
                'color': 'FFA500',  # Оранжевый
                'header': 'Отсутствующие сотрудники',
                'columns': ['ID', 'ФИО', 'Статус']
            }
        }

        # Создание листов
        for sheet_name, config in sheets_config.items():
            ws = wb.create_sheet(sheet_name)

            # Заголовки столбцов
            ws.append(config['columns'])

            # Настройка стиля заголовков
            header_font = Font(bold=True, color=config['color'])
            for cell in ws[1]:
                cell.font = header_font

            # Добавление данных
            for row in config['data']:
                ws.append(row)

            # Автоматическая настройка ширины столбцов
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранение и закрытие файла
        wb.save(output_path)
        wb.close()
        return output_path


if __name__ == '__main__':
    # Точка входа в приложение
    app = QApplication(sys.argv)
    window = ExcelComparator()
    window.show()
    sys.exit(app.exec())
